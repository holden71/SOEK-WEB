"""
Excel file processing endpoints
"""
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from typing import Optional
import openpyxl
import io
import re

router = APIRouter(prefix="/api", tags=["excel"])


@router.post("/analyze-excel")
async def analyze_excel(
    file: UploadFile = File(...),
    filter_percentage_only: bool = False
):
    """Analyze Excel file and return sheet information"""
    try:
        # Read file content
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        sheets = []
        for sheet_name in workbook.sheetnames:
            # If filter enabled, only include sheets with percentage names
            if filter_percentage_only:
                # Check if sheet name matches percentage format (4%, 0.2%, 1,2%, etc.)
                if not re.match(r'^\d+([.,]\d+)?%$', sheet_name.strip()):
                    continue

            sheet = workbook[sheet_name]
            # Count rows with data
            row_count = 0
            for row in sheet.iter_rows():
                if any(cell.value is not None for cell in row):
                    row_count += 1

            sheets.append({
                "name": sheet_name,
                "rows": row_count
            })

        workbook.close()

        return {"sheets": sheets}

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to analyze Excel file: {str(e)}"
        )


@router.post("/extract-sheet-data")
async def extract_sheet_data(
    file: UploadFile = File(...),
    sheet_name: str = Form(...)
):
    """Extract data from specific Excel sheet"""
    try:
        # Read file content
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

        # Check if sheet exists
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{sheet_name}' not found in Excel file"
            )

        sheet = workbook[sheet_name]

        # Get headers from first row - skip empty leading columns
        headers = []
        header_indices = []  # Track which column indices have headers

        # Read first 5 rows to find where headers are
        first_rows = list(sheet.iter_rows(min_row=1, max_row=5, values_only=True))

        if not first_rows:
            raise HTTPException(
                status_code=400,
                detail="Sheet is empty"
            )

        # Try to find the header row (first non-empty row with text values)
        header_row_idx = None
        first_row = None

        for i, row in enumerate(first_rows, start=1):
            if row and any(cell is not None and not isinstance(cell, (int, float)) for cell in row):
                # Found a row with text values
                header_row_idx = i
                first_row = row
                break

        if not first_row:
            raise HTTPException(
                status_code=400,
                detail="No header row found in first 5 rows"
            )

        for idx, cell in enumerate(first_row):
            if cell is not None and str(cell).strip():
                headers.append(str(cell).strip())
                header_indices.append(idx)

        if not headers:
            raise HTTPException(
                status_code=400,
                detail="No valid column headers found in first row"
            )

        # Extract data only from columns that have headers
        data = {header: [] for header in headers}

        # Start reading data from the row after headers
        data_start_row = header_row_idx + 1

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            for header_idx, col_idx in enumerate(header_indices):
                if col_idx < len(row):
                    cell = row[col_idx]
                    # Convert to appropriate type
                    if cell is None:
                        data[headers[header_idx]].append(None)
                    elif isinstance(cell, (int, float)):
                        data[headers[header_idx]].append(float(cell))
                    else:
                        data[headers[header_idx]].append(str(cell))

        workbook.close()

        return data

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Failed to extract data from sheet: {str(e)}"
        )
