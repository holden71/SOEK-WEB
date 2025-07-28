#!/usr/bin/env -S uv run

import io
import os
import re
import tempfile
from contextlib import asynccontextmanager
from typing import List

import openpyxl
from db import DbSessionDep, DbSessionManager
from fastapi import (Body, Depends, FastAPI, File, Form, HTTPException, Query,
                     UploadFile)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, JSONResponse
from models import Plant, SearchData, SetAccelProcedureParams, SetAccelProcedureResult, Term, Unit, FindReqAccelSetParams, FindReqAccelSetResult, ClearAccelSetParams, ClearAccelSetResult, SpectralDataResult, SaveAnalysisResultParams, SaveAnalysisResultResponse, SaveStressInputsParams, SaveStressInputsResponse, SaveKResultsParams, SaveKResultsResponse
from pydantic import BaseModel
from settings import settings
from sqlalchemy import inspect, text


@asynccontextmanager
async def lifespan(_app: FastAPI):
    DbSessionManager.initialize()
    yield
    DbSessionManager.dispose()
    print("Database connection closed")


app = FastAPI(lifespan=lifespan)

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins
    allow_credentials=True,
    allow_methods=["*"],  # Allows all methods
    allow_headers=["*"],  # Allows all headers
)

@app.get("/api/plants", response_model=List[Plant])
async def get_plants(db: DbSessionDep):
    """Get all plants from UNS_PLANTS table"""
    result = db.execute(text("SELECT NAME, PLANT_ID FROM UNS_PLANTS ORDER BY NAME"))
    plants = [Plant(name=row[0], plant_id=row[1]) for row in result]
    return plants

@app.get("/api/units", response_model=List[Unit])
async def get_units(
    db: DbSessionDep,
    plant_id: int = Query(..., description="ID of the plant to get units for")
):
    """Get all units for a specific plant from UNS_UNITS table"""
    result = db.execute(
        text("SELECT NAME, UNIT_ID FROM UNS_UNITS WHERE PLANT_ID = :plant_id ORDER BY NAME"),
        {"plant_id": plant_id}
    )
    units = [Unit(name=row[0], unit_id=row[1]) for row in result]
    return units

@app.get("/api/terms", response_model=List[Term])
async def get_terms(
    db: DbSessionDep,
    plant_id: int = Query(..., description="ID of the plant"),
    unit_id: int = Query(..., description="ID of the unit")
):
    """Get all terms for a specific plant and unit from SRT_TERMS_LOC table"""
    result = db.execute(
        text("""
            SELECT T_NAME, T_ID 
            FROM SRT_TERMS_LOC 
            WHERE PLANT_ID = :plant_id 
            AND UNIT_ID = :unit_id 
            ORDER BY T_NAME
        """),
        {"plant_id": plant_id, "unit_id": unit_id}
    )
    terms = [Term(name=row[0], term_id=row[1]) for row in result]
    return terms

@app.get("/api/search", response_model=List[SearchData])
async def search_data(
    db: DbSessionDep,
    plant_id: int = Query(..., description="ID of the plant"),
    unit_id: int = Query(..., description="ID of the unit"),
    t_id: int = Query(..., description="Term ID (EKLIST_ID)")
):
    """Search data from SRTN_EK_SEISM_DATA table with dynamic column selection"""
    # First, get all columns from the table
    inspector = inspect(db.get_bind())
    columns = inspector.get_columns('SRTN_EK_SEISM_DATA')
    column_names = [col['name'] for col in columns]
    
    # Build the dynamic query
    query = text(f"""
        SELECT {', '.join(column_names)}
        FROM SRTN_EK_SEISM_DATA
        WHERE PLANT_ID = :plant_id
        AND UNIT_ID = :unit_id
        AND EKLIST_ID = :t_id
    """)
    
    result = db.execute(query, {
        "plant_id": plant_id,
        "unit_id": unit_id,
        "t_id": t_id
    })
    
    # Convert each row to a dictionary with column names as keys
    search_results = []
    for row in result:
        row_dict = {column_names[i]: value for i, value in enumerate(row)}
        search_results.append(SearchData(data=row_dict))
    
    return search_results

class LocationCheck(BaseModel):
    plant_id: int
    unit_id: int
    building: str
    room: str

@app.post("/api/check-location")
async def check_location(
    db: DbSessionDep,
    location: LocationCheck = Body(...)
):
    """Check if building and room exist in SRTN_EK_SEISM_DATA table"""
    result = db.execute(
        text("""
            SELECT COUNT(*) 
            FROM SRTN_EK_SEISM_DATA 
            WHERE PLANT_ID = :plant_id 
            AND UNIT_ID = :unit_id 
            AND BUILDING = :building 
            AND ROOM = :room
        """),
        {
            "plant_id": location.plant_id,
            "unit_id": location.unit_id,
            "building": location.building,
            "room": location.room
        }
    )
    count = result.scalar()
    return {"exists": count > 0}

class BuildingCheck(BaseModel):
    plant_id: int
    unit_id: int
    building: str

@app.post("/api/check-building")
async def check_building(
    db: DbSessionDep,
    data: BuildingCheck = Body(...)
):
    """Check if building exists in SRTN_EK_SEISM_DATA table for given plant and unit"""
    result = db.execute(
        text("""
            SELECT COUNT(*)
            FROM SRTN_EK_SEISM_DATA
            WHERE PLANT_ID = :plant_id
            AND UNIT_ID = :unit_id
            AND BUILDING = :building
        """),
        {
            "plant_id": data.plant_id,
            "unit_id": data.unit_id,
            "building": data.building
        }
    )
    count = result.scalar()
    return {"exists": count > 0}

@app.post("/api/analyze-excel")
async def analyze_excel(
    file: UploadFile = File(...),
    filter_percentage_only: bool = Query(False, description="Filter sheets to include only percentage names")
):
    """
    Analyze Excel file and return information about sheets.
    Optionally filter to include only sheets with percentage names (like 4%, 0.2%, 1,2%)
    """
    if not file.filename.endswith(('.xls', '.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xls, .xlsx, .xlsm)")
    
    try:
        # Read file directly into memory
        contents = await file.read()
        file_object = io.BytesIO(contents)
        
        # Load workbook from memory buffer
        workbook = openpyxl.load_workbook(file_object, read_only=True, data_only=True)
        
        # Get sheet names
        sheet_names = workbook.sheetnames
        
        # Regex pattern for percentage names (e.g., 4%, 0.2%, 1,2%)
        # This matches numbers with optional decimal parts (using . or , as separator) followed by %
        percentage_pattern = re.compile(r'^[0-9]+([.,][0-9]+)?%$')
        
        # Get basic info about each sheet
        sheets_info = []
        for sheet_name in sheet_names:
            # Skip non-percentage sheets if filter is enabled
            if filter_percentage_only and not percentage_pattern.match(sheet_name.strip()):
                continue
                
            sheet = workbook[sheet_name]
            # Get dimensions
            try:
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row > 0 and max_col > 0:
                    sheets_info.append({
                        "name": sheet_name,
                        "rows": max_row,
                        "columns": max_col
                    })
                else:
                    sheets_info.append({
                        "name": sheet_name,
                        "rows": 0,
                        "columns": 0
                    })
            except Exception as sheet_error:
                # Add sheet with error info
                sheets_info.append({
                    "name": sheet_name,
                    "rows": 0,
                    "columns": 0,
                    "error": str(sheet_error)
                })
        
        # Make sure to close workbook and file objects
        workbook.close()
        file_object.close()
        
        return {"sheets": sheets_info}
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing Excel file: {str(e)}")

@app.post("/api/extract-sheet-data")
async def extract_sheet_data(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
):
    """
    Extract data from a specific sheet in an Excel file.
    Returns a dictionary with sheet_name and column contents.
    """
    if not file.filename.endswith(('.xls', '.xlsx', '.xlsm')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xls, .xlsx, .xlsm)")
    
    try:
        # Read file directly into memory
        contents = await file.read()
        file_object = io.BytesIO(contents)
        
        # Load workbook from memory buffer
        workbook = openpyxl.load_workbook(file_object, read_only=True, data_only=True)
        
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found in the Excel file")
        
        sheet = workbook[sheet_name]
        
        # Get column headers (assuming first row contains headers)
        headers = []
        max_col = sheet.max_column
        
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value))
        
        # Initialize result dictionary with sheet name
        result = {
            'demp': sheet_name
        }
        
        # Extract column data
        for col_idx, header in enumerate(headers, start=1):
            column_data = []
            for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (skip headers)
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    column_data.append(str(cell_value))
            
            # Use column number as key if header is empty
            col_key = f"column_{col_idx}" if not header.strip() else header
            result[col_key] = column_data
        
        # Close workbook and file objects
        workbook.close()
        file_object.close()
        
        return result
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting sheet data: {str(e)}")

class AccelDataItem(BaseModel):
    dempf: float | None = None  # Make dempf nullable
    data: dict

class AccelData(BaseModel):
    plant_id: int
    unit_id: int
    building: str
    room: str = None
    lev: float | None = None   # Add LEV field
    lev1: float | None = None  # Using more explicit Union type for None
    lev2: float | None = None  # Using more explicit Union type for None
    pga: float | None = None   # Using more explicit Union type for None
    calc_type: str
    set_type: str = "ВИМОГИ"   # Default to "ВИМОГИ" but allow custom values
    sheets: dict[str, AccelDataItem]

@app.post("/api/save-accel-data")
async def save_accel_data(
    db: DbSessionDep,
    data: AccelData = Body(...)
):
    """
    Save acceleration data to the database tables: 
    SRTN_ACCEL_SET, SRTN_ACCEL_PLOT, and SRTN_ACCEL_POINT
    """
    try:
        # Debug: Get actual column names from SRTN_ACCEL_SET table
        inspector = inspect(db.get_bind())
        columns = inspector.get_columns('SRTN_ACCEL_SET')
        column_names = [col['name'] for col in columns]
        print("Available columns in SRTN_ACCEL_SET:", column_names)
        
        # Check if PGA column exists and note its exact name
        pga_column = None
        for col in column_names:
            if col.upper() == 'PGA' or col.upper().startswith('PGA_'):
                pga_column = col
                print(f"Found PGA column: {pga_column}")
                break

        # Get plant and unit names
        plant_result = db.execute(
            text("SELECT NAME FROM UNS_PLANTS WHERE PLANT_ID = :plant_id"),
            {"plant_id": data.plant_id}
        )
        plant_name = plant_result.scalar()

        unit_result = db.execute(
            text("SELECT NAME FROM UNS_UNITS WHERE UNIT_ID = :unit_id"),
            {"unit_id": data.unit_id}
        )
        unit_name = unit_result.scalar()

        if not plant_name or not unit_name:
            raise HTTPException(status_code=400, detail="Invalid plant or unit ID")

        # Store created record IDs for the response
        created_records = {
            "sets": [],
            "plots": [],
            "points": 0,
            "mrz_set_id": None,  # Store MRZ set ID for procedure
            "pz_set_id": None    # Store PZ set ID for procedure
        }

        # Process each sheet
        for sheet_name, sheet_data in data.sheets.items():
            dempf = sheet_data.dempf  # This can now be None
            sheet_columns = sheet_data.data

            # Get frequency column
            freq_column = None
            for col_name in sheet_columns.keys():
                if col_name.lower() in ["частота", "частота, гц", "частота,гц", "частота гц", "част", "frequency", "freq", "hz"]:
                    freq_column = col_name
                    break

            if not freq_column:
                # Assume first column is frequency if no clear frequency column found
                columns = list(sheet_columns.keys())
                if columns:
                    freq_column = columns[0]
                else:
                    continue  # Skip this sheet if no data

            # Filter to only keep relevant columns
            valid_spectrum_types = ["МРЗ", "ПЗ"]
            valid_axes = ["x", "y", "z"]
            
            # Find all valid columns that match our pattern (МРЗ_x, МРЗ_y, etc.)
            relevant_columns = [freq_column]  # Always include frequency column
            
            for column_name in sheet_columns.keys():
                if "_" in column_name:
                    parts = column_name.split("_")
                    if len(parts) == 2 and parts[0] in valid_spectrum_types and parts[1].lower() in valid_axes:
                        relevant_columns.append(column_name)
            
            # Only process spectrum types that exist in the relevant columns
            spectrum_types = set()
            for column_name in relevant_columns:
                if "_" in column_name:
                    spectrum_type = column_name.split("_")[0]
                    if spectrum_type in valid_spectrum_types:
                        spectrum_types.add(spectrum_type)

            # Process each spectrum type (МРЗ, ПЗ)
            for spectrum_type in spectrum_types:
                # First create a set record in SRTN_ACCEL_SET
                
                # Use LEV from the data instead of hardcoding to null
                lev = data.lev
                
                set_params = {
                    "set_type": data.set_type,
                    "building": data.building,
                    "room": data.room,
                    "lev": lev,
                    "lev1": data.lev1,
                    "lev2": data.lev2,
                    "dempf": dempf,
                    "plant_id": data.plant_id,
                    "plant_name": plant_name,
                    "unit_id": data.unit_id,
                    "unit_name": unit_name,
                    "spectr_type": spectrum_type,  # МРЗ/ПЗ
                    "calc_type": data.calc_type  # ДЕТЕРМІНИСТИЧНИЙ/ІМОВІРНІСНИЙ
                }
                
                db.execute(text("""
                    INSERT INTO SRTN_ACCEL_SET(SET_TYPE, BUILDING, ROOM, LEV, LEV1, LEV2, DEMPF, 
                    PLANT_ID, PLANT_NAME, UNIT_ID, UNIT_NAME, SPECTR_EARTHQ_TYPE, CALC_TYPE) 
                    VALUES(:set_type, :building, :room, :lev, :lev1, :lev2, :dempf, :plant_id, 
                    :plant_name, :unit_id, :unit_name, :spectr_type, :calc_type)
                """), set_params)
                
                # Get the last inserted ID using a separate query
                id_query = text("""
                    SELECT ACCEL_SET_ID FROM (
                        SELECT ACCEL_SET_ID FROM SRTN_ACCEL_SET 
                        WHERE PLANT_ID = :plant_id 
                        AND UNIT_ID = :unit_id 
                        AND BUILDING = :building
                        AND ((:dempf IS NULL AND DEMPF IS NULL) OR DEMPF = :dempf)
                        AND SPECTR_EARTHQ_TYPE = :spectr_type
                        ORDER BY ACCEL_SET_ID DESC
                    ) WHERE ROWNUM = 1
                """)
                
                # Debug the parameters to check NULL values
                print(f"Query parameters: plant_id={data.plant_id}, unit_id={data.unit_id}, building={data.building}, dempf={dempf}, spectr_type={spectrum_type}")
                
                id_result = db.execute(id_query, {
                    "plant_id": data.plant_id,
                    "unit_id": data.unit_id,
                    "building": data.building,
                    "dempf": dempf,
                    "spectr_type": spectrum_type
                })
                
                set_id = id_result.scalar()
                created_records["sets"].append(set_id)
                
                # Store specific set IDs for procedure parameters
                if spectrum_type == "МРЗ":
                    created_records["mrz_set_id"] = set_id
                    print(f"Stored MRZ set ID: {set_id}")
                elif spectrum_type == "ПЗ":
                    created_records["pz_set_id"] = set_id
                    print(f"Stored PZ set ID: {set_id}")
                
                # Only update PGA if it's not None (for Import.jsx)
                # For Main.jsx imports, data.pga will be null so this section will be skipped
                if data.pga is not None:
                    try:
                        # Try updating with column name we discovered
                        if pga_column:
                            update_pga_query = text(f"""
                                UPDATE SRTN_ACCEL_SET
                                SET {pga_column} = :pga_value
                                WHERE ACCEL_SET_ID = :set_id
                            """)
                            
                            db.execute(update_pga_query, {
                                "pga_value": data.pga,
                                "set_id": set_id
                            })
                        else:
                            # Try alternative column names if pga_column not found
                            try_columns = ['FGA', 'PGA']  # Try FGA as alternative
                            for col_name in try_columns:
                                try:
                                    update_query = text(f"""
                                        UPDATE SRTN_ACCEL_SET
                                        SET {col_name} = :pga_value
                                        WHERE ACCEL_SET_ID = :set_id
                                    """)
                                    
                                    db.execute(update_query, {
                                        "pga_value": data.pga,
                                        "set_id": set_id
                                    })
                                    print(f"Successfully updated {col_name} column")
                                    break
                                except Exception as err:
                                    print(f"Failed to update with column {col_name}: {err}")
                    except Exception as pga_error:
                        print(f"Error updating PGA value: {pga_error}")
                
                # Create plot records for X, Y, Z axes
                plot_ids = {}
                for axis in ["x", "y", "z"]:
                    column_name = f"{spectrum_type}_{axis}"
                    if column_name not in sheet_columns:
                        continue
                        
                    modified_plot_query = text("""
                        INSERT INTO SRTN_ACCEL_PLOT (AXIS, NAME)
                        VALUES (:axis, :name)
                    """)
                    
                    plot_params = {
                        "axis": axis.upper(),
                        "name": column_name
                    }
                    
                    db.execute(modified_plot_query, plot_params)
                    
                    # Get the last inserted plot ID
                    plot_id_query = text("""
                        SELECT PLOT_ID FROM (
                            SELECT PLOT_ID FROM SRTN_ACCEL_PLOT 
                            WHERE AXIS = :axis AND NAME = :name
                            ORDER BY PLOT_ID DESC
                        ) WHERE ROWNUM = 1
                    """)
                    
                    plot_id_result = db.execute(plot_id_query, plot_params)
                    plot_id = plot_id_result.scalar()
                    
                    plot_ids[axis] = plot_id
                    created_records["plots"].append(plot_id)
                
                # Update set record with plot IDs
                update_set_query = text("""
                    UPDATE SRTN_ACCEL_SET
                    SET X_PLOT_ID = :x_plot_id, Y_PLOT_ID = :y_plot_id, Z_PLOT_ID = :z_plot_id
                    WHERE ACCEL_SET_ID = :set_id
                """)
                
                update_set_params = {
                    "x_plot_id": plot_ids.get("x"),
                    "y_plot_id": plot_ids.get("y"),
                    "z_plot_id": plot_ids.get("z"),
                    "set_id": set_id
                }
                
                db.execute(update_set_query, update_set_params)
                
                # Insert frequency-acceleration points for each axis
                for axis in ["x", "y", "z"]:
                    column_name = f"{spectrum_type}_{axis}"
                    if column_name not in sheet_columns or column_name not in relevant_columns or axis not in plot_ids:
                        continue
                        
                    plot_id = plot_ids[axis]
                    frequencies = sheet_columns.get(freq_column, [])
                    accelerations = sheet_columns.get(column_name, [])
                    
                    # Make sure we have equal number of values
                    num_points = min(len(frequencies), len(accelerations))
                    
                    for i in range(num_points):
                        try:
                            freq = float(frequencies[i])
                            accel = float(accelerations[i])
                            
                            point_query = text("""
                                INSERT INTO SRTN_ACCEL_POINT (FREQ, ACCEL, PLOT_ID)
                                VALUES (:freq, :accel, :plot_id)
                            """)
                            
                            point_params = {
                                "freq": freq,
                                "accel": accel,
                                "plot_id": plot_id
                            }
                            
                            db.execute(point_query, point_params)
                            created_records["points"] += 1
                        except (ValueError, TypeError):
                            # Skip invalid numeric values
                            continue
        
        # Commit all changes
        db.commit()
        
        # Get final mrz_set_id and pz_set_id values for return
        mrz_set_id = created_records.get("mrz_set_id")
        pz_set_id = created_records.get("pz_set_id")
        
        return {
            "success": True,
            "message": "Data saved successfully",
            "created": created_records,
            "mrz_set_id": mrz_set_id,
            "pz_set_id": pz_set_id
        }
        
    except Exception as e:
        # Rollback in case of error
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error saving acceleration data: {str(e)}")

@app.post("/api/execute-set-all-ek-accel-set", response_model=SetAccelProcedureResult)
async def execute_set_all_ek_accel_set(
    db: DbSessionDep,
    params: SetAccelProcedureParams = Body(...)
):
    """
    Execute the SET_ALL_EK_ACCEL_SET stored procedure
    This procedure sets acceleration data for an element or all elements of the same type
    """
    try:
        # Prepare the parameters
        ek_id = params.ek_id
        set_mrz = params.set_mrz
        set_pz = params.set_pz
        can_overwrite = params.can_overwrite
        do_for_all = params.do_for_all
        clear_sets = params.clear_sets
        
        # Log input parameters
        print(f"Executing procedure with: EK_ID={ek_id}, SET_MRZ={set_mrz}, SET_PZ={set_pz}, CAN_OVERWRITE={can_overwrite}, DO_FOR_ALL={do_for_all}, CLEAR_SETS={clear_sets}")
        
        # Execute procedure using cursor.callproc (most reliable approach)
        conn = db.connection().connection
        cursor = conn.cursor()
        
        # Create output variables for all 7 output parameters
        done_for_id_mrz = cursor.var(int)
        done_for_id_pz = cursor.var(int)
        done_for_all_mrz = cursor.var(int)
        done_for_all_pz = cursor.var(int)
        total_ek = cursor.var(int)
        processed_mrz = cursor.var(int)
        processed_pz = cursor.var(int)
        
        # Call the procedure with updated signature
        cursor.callproc('SET_ALL_EK_ACCEL_SET', [
            ek_id, set_mrz, set_pz, can_overwrite, do_for_all, clear_sets,
            done_for_id_mrz, done_for_id_pz, done_for_all_mrz, done_for_all_pz,
            total_ek, processed_mrz, processed_pz
        ])
        
        # Get the output values
        done_for_id_mrz_value = done_for_id_mrz.getvalue()
        done_for_id_pz_value = done_for_id_pz.getvalue()
        done_for_all_mrz_value = done_for_all_mrz.getvalue() if do_for_all else None
        done_for_all_pz_value = done_for_all_pz.getvalue() if do_for_all else None
        total_ek_value = total_ek.getvalue()
        processed_mrz_value = processed_mrz.getvalue()
        processed_pz_value = processed_pz.getvalue()
        
        # Log results
        print(f"Procedure results:")
        print(f"  done_for_id_mrz={done_for_id_mrz_value}, done_for_id_pz={done_for_id_pz_value}")
        print(f"  done_for_all_mrz={done_for_all_mrz_value}, done_for_all_pz={done_for_all_pz_value}")
        print(f"  total_ek={total_ek_value}, processed_mrz={processed_mrz_value}, processed_pz={processed_pz_value}")
        print(f"Element MRZ update: {'Success' if done_for_id_mrz_value == 1 else 'Failed'}")
        print(f"Element PZ update: {'Success' if done_for_id_pz_value == 1 else 'Failed'}")
        if do_for_all:
            print(f"All elements MRZ update: {'Success' if done_for_all_mrz_value == 1 else 'Failed'}")
            print(f"All elements PZ update: {'Success' if done_for_all_pz_value == 1 else 'Failed'}")
        
        # Commit changes
        db.commit()
        
        # Return values directly from procedure
        return SetAccelProcedureResult(
            done_for_id_mrz=done_for_id_mrz_value,
            done_for_id_pz=done_for_id_pz_value,
            done_for_all_mrz=done_for_all_mrz_value,
            done_for_all_pz=done_for_all_pz_value,
            total_ek=total_ek_value,
            processed_mrz=processed_mrz_value,
            processed_pz=processed_pz_value
        )
        
    except Exception as e:
        # Rollback in case of error
        db.rollback()
        print(f"Procedure execution failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error executing SET_ALL_EK_ACCEL_SET procedure: {str(e)}")

@app.post("/api/find-req-accel-set", response_model=FindReqAccelSetResult)
async def find_req_accel_set(
    db: DbSessionDep,
    params: FindReqAccelSetParams = Body(...)
):
    """
    Execute the FIND_REQ_ACCEL_SET stored procedure to check for existing acceleration data
    """
    try:
        # Prepare the parameters
        plant_id = params.plant_id
        unit_id = params.unit_id
        building = params.building
        room = params.room
        lev1 = params.lev1
        lev2 = params.lev2
        earthq_type = params.earthq_type
        calc_type = params.calc_type
        set_type = params.set_type
        dempf = params.dempf
        
        # Log input parameters
        print(f"Executing FIND_REQ_ACCEL_SET with: PLANT_ID={plant_id}, UNIT_ID={unit_id}, BUILDING={building}, ROOM={room}, LEV1={lev1}, LEV2={lev2}, EARTHQ_TYPE={earthq_type}, CALC_TYPE={calc_type}, SET_TYPE={set_type}, DEMPF={dempf}")
        
        # Execute procedure using cursor.callproc
        conn = db.connection().connection
        cursor = conn.cursor()
        
        # Create output variables
        set_id_out = cursor.var(int)
        found_ek_out = cursor.var(int)
        
        # Call the procedure
        cursor.callproc('FIND_REQ_ACCEL_SET', [
            plant_id, unit_id, building, room, lev1, lev2, earthq_type, calc_type, set_type, dempf,
            set_id_out, found_ek_out
        ])
        
        # Get the output values
        set_id_value = set_id_out.getvalue()
        found_ek_value = found_ek_out.getvalue()
        
        # Log results
        print(f"FIND_REQ_ACCEL_SET results: set_id={set_id_value}, found_ek={found_ek_value}")
        
        return FindReqAccelSetResult(
            set_id=set_id_value,
            found_ek=found_ek_value
        )
        
    except Exception as e:
        print(f"FIND_REQ_ACCEL_SET execution failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error executing FIND_REQ_ACCEL_SET procedure: {str(e)}")

@app.post("/api/clear-accel-set-arrays", response_model=ClearAccelSetResult)
async def clear_accel_set_arrays(
    db: DbSessionDep,
    params: ClearAccelSetParams = Body(...)
):
    """
    Execute the CLEAR_ACCEL_CET_ARRAYS stored procedure to clear existing acceleration data
    """
    try:
        # Prepare the parameters
        set_id = params.set_id
        
        # Log input parameters
        print(f"Executing CLEAR_ACCEL_CET_ARRAYS with: SET_ID={set_id}")
        
        # Execute procedure using cursor.callproc
        conn = db.connection().connection
        cursor = conn.cursor()
        
        # Create output variable
        clear_result_out = cursor.var(str)
        
        # Call the procedure
        cursor.callproc('CLEAR_ACCEL_CET_ARRAYS', [
            set_id, clear_result_out
        ])
        
        # Get the output value
        clear_result_value = clear_result_out.getvalue()
        
        # Log results
        print(f"CLEAR_ACCEL_CET_ARRAYS results: clear_result={clear_result_value}")
        
        # Commit the changes
        db.commit()
        
        return ClearAccelSetResult(
            clear_result=clear_result_value
        )
        
    except Exception as e:
        # Rollback in case of error
        db.rollback()
        print(f"CLEAR_ACCEL_CET_ARRAYS execution failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error executing CLEAR_ACCEL_CET_ARRAYS procedure: {str(e)}")

@app.get("/api/spectral-data", response_model=SpectralDataResult)
async def get_spectral_data(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element"),
    calc_type: str = Query(..., description="Calculation type (ДЕТЕРМІНИСТИЧНИЙ/ІМОВІРНІСНИЙ)"),
    spectrum_type: str = Query(..., description="Spectrum type (МРЗ/ПЗ)")
):
    """
    Get spectral data for a specific element.
    
    This endpoint:
    1. Finds the element in SRTN_EK_SEISM_DATA
    2. Looks for corresponding acceleration sets in SRTN_ACCEL_SET
    3. Retrieves plot data from SRTN_ACCEL_PLOT and SRTN_ACCEL_POINT
    4. Returns frequency and acceleration data for X, Y, Z axes
    """
    try:
        # Step 1: Get element data to find building, room, and other parameters
        element_query = text("""
            SELECT BUILDING, ROOM, LEV, PLANT_ID, UNIT_ID
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        element_result = db.execute(element_query, {"ek_id": ek_id})
        element_row = element_result.fetchone()
        
        if not element_row:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        building = element_row[0]
        room = element_row[1]
        lev = element_row[2]
        plant_id = element_row[3]
        unit_id = element_row[4]
        
        # Step 2: Find acceleration set matching the criteria
        set_query = text("""
            SELECT ACCEL_SET_ID, X_PLOT_ID, Y_PLOT_ID, Z_PLOT_ID
            FROM SRTN_ACCEL_SET
            WHERE PLANT_ID = :plant_id
            AND UNIT_ID = :unit_id
            AND BUILDING = :building
            AND (ROOM = :room OR (ROOM IS NULL AND :room IS NULL))
            AND (LEV = :lev OR (LEV IS NULL AND :lev IS NULL))
            AND SPECTR_EARTHQ_TYPE = :spectrum_type
            AND CALC_TYPE = :calc_type
            AND SET_TYPE = 'ХАРАКТЕРИСТИКИ'
            ORDER BY ACCEL_SET_ID DESC
        """)
        
        set_result = db.execute(set_query, {
            "plant_id": plant_id,
            "unit_id": unit_id,
            "building": building,
            "room": room,
            "lev": lev,
            "spectrum_type": spectrum_type,
            "calc_type": calc_type
        })
        
        set_row = set_result.fetchone()
        
        if not set_row:
            # Return empty data structure if no set found
            return SpectralDataResult(frequency=[])
        
        accel_set_id = set_row[0]
        x_plot_id = set_row[1]
        y_plot_id = set_row[2]
        z_plot_id = set_row[3]
        
        # Step 3: Get spectral data for each axis
        result_data = {"frequency": []}
        
        # Helper function to get plot data
        def get_plot_data(plot_id, axis_name):
            if not plot_id:
                return []
            
            point_query = text("""
                SELECT FREQ, ACCEL
                FROM SRTN_ACCEL_POINT
                WHERE PLOT_ID = :plot_id
                ORDER BY FREQ
            """)
            
            point_result = db.execute(point_query, {"plot_id": plot_id})
            points = point_result.fetchall()
            
            frequencies = []
            accelerations = []
            
            for point in points:
                freq = float(point[0]) if point[0] is not None else 0.0
                accel = float(point[1]) if point[1] is not None else 0.0
                frequencies.append(freq)
                accelerations.append(accel)
            
            return frequencies, accelerations
        
        # Get data for each axis
        x_freq, x_accel = get_plot_data(x_plot_id, 'x')
        y_freq, y_accel = get_plot_data(y_plot_id, 'y')
        z_freq, z_accel = get_plot_data(z_plot_id, 'z')
        
        # Use the longest frequency array as the base
        all_frequencies = [x_freq, y_freq, z_freq]
        base_freq = max(all_frequencies, key=len) if any(all_frequencies) else []
        
        # Build response
        response_data = {
            "frequency": base_freq
        }
        
        # Add acceleration data based on spectrum type
        if spectrum_type == "МРЗ":
            response_data["mrz_x"] = x_accel if x_accel else None
            response_data["mrz_y"] = y_accel if y_accel else None
            response_data["mrz_z"] = z_accel if z_accel else None
        elif spectrum_type == "ПЗ":
            response_data["pz_x"] = x_accel if x_accel else None
            response_data["pz_y"] = y_accel if y_accel else None
            response_data["pz_z"] = z_accel if z_accel else None
        
        return SpectralDataResult(**response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting spectral data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving spectral data: {str(e)}")

@app.get("/api/seism-requirements", response_model=SpectralDataResult)
async def get_seism_requirements(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element"),
    dempf: float = Query(..., description="Damping factor"),
    spectr_earthq_type: str = Query("МРЗ", description="Spectrum earthquake type"),
    calc_type: str = Query("ДЕТЕРМІНИСТИЧНИЙ", description="Calculation type")
):
    """
    Get seismic requirements data using the GET_SEISM_REQUIREMENTS procedure.
    
    This endpoint:
    1. Calls GET_SEISM_REQUIREMENTS procedure to get SET_ID
    2. If SET_ID is found, retrieves spectral data for requirements
    3. Returns frequency and acceleration data for X, Y, Z axes
    """
    try:
        # Execute GET_SEISM_REQUIREMENTS procedure
        conn = db.connection().connection
        cursor = conn.cursor()
        
        # Create output variable for SET_ID
        set_id_out = cursor.var(int)
        
        # Call the procedure
        cursor.callproc('GET_SEISM_REQUIREMENTS', [
            ek_id, dempf, spectr_earthq_type, calc_type, set_id_out
        ])
        
        # Get the output SET_ID
        set_id = set_id_out.getvalue()
        
        print(f"GET_SEISM_REQUIREMENTS results: SET_ID={set_id}")
        
        if not set_id:
            # Return empty data structure if no requirements found
            return SpectralDataResult(frequency=[])
        
        # Get acceleration set data
        set_query = text("""
            SELECT ACCEL_SET_ID, X_PLOT_ID, Y_PLOT_ID, Z_PLOT_ID
            FROM SRTN_ACCEL_SET
            WHERE ACCEL_SET_ID = :set_id
        """)
        
        set_result = db.execute(set_query, {"set_id": set_id})
        set_row = set_result.fetchone()
        
        if not set_row:
            return SpectralDataResult(frequency=[])
        
        x_plot_id = set_row[1]
        y_plot_id = set_row[2]
        z_plot_id = set_row[3]
        
        # Helper function to get plot data
        def get_plot_data(plot_id):
            if not plot_id:
                return [], []
            
            point_query = text("""
                SELECT FREQ, ACCEL
                FROM SRTN_ACCEL_POINT
                WHERE PLOT_ID = :plot_id
                ORDER BY FREQ
            """)
            
            point_result = db.execute(point_query, {"plot_id": plot_id})
            points = point_result.fetchall()
            
            frequencies = []
            accelerations = []
            
            for point in points:
                freq = float(point[0]) if point[0] is not None else 0.0
                accel = float(point[1]) if point[1] is not None else 0.0
                frequencies.append(freq)
                accelerations.append(accel)
            
            return frequencies, accelerations
        
        # Get data for each axis
        x_freq, x_accel = get_plot_data(x_plot_id)
        y_freq, y_accel = get_plot_data(y_plot_id)
        z_freq, z_accel = get_plot_data(z_plot_id)
        
        # Use the longest frequency array as the base
        all_frequencies = [x_freq, y_freq, z_freq]
        base_freq = max(all_frequencies, key=len) if any(all_frequencies) else []
        
        # Build response
        response_data = {
            "frequency": base_freq
        }
        
        # Add acceleration data based on spectrum type
        if spectr_earthq_type == "МРЗ":
            response_data["mrz_x"] = x_accel if x_accel else None
            response_data["mrz_y"] = y_accel if y_accel else None
            response_data["mrz_z"] = z_accel if z_accel else None
        elif spectr_earthq_type == "ПЗ":
            response_data["pz_x"] = x_accel if x_accel else None
            response_data["pz_y"] = y_accel if y_accel else None
            response_data["pz_z"] = z_accel if z_accel else None
        
        return SpectralDataResult(**response_data)
        
    except Exception as e:
        print(f"Error getting seism requirements: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving seism requirements: {str(e)}")

@app.post("/api/save-analysis-result", response_model=SaveAnalysisResultResponse)
async def save_analysis_result(
    db: DbSessionDep,
    params: SaveAnalysisResultParams = Body(...)
):
    """
    Save analysis results (m1, m2) to SRTN_EK_SEISM_DATA table.
    
    This endpoint:
    1. Updates the SRTN_EK_SEISM_DATA table with calculated m1 and m2 values
    2. Uses different column names based on spectrum type:
       - МРЗ: M1_MRZ, M2_MRZ
       - ПЗ: M1_PZ, M2_PZ
    """
    try:
        # Validate spectrum type
        if params.spectrum_type not in ["МРЗ", "ПЗ"]:
            raise HTTPException(status_code=400, detail="Invalid spectrum type. Must be МРЗ or ПЗ")
        
        # Determine column names based on spectrum type
        if params.spectrum_type == "МРЗ":
            m1_column = "M1_MRZ"
            m2_column = "M2_MRZ"
        else:  # ПЗ
            m1_column = "M1_PZ"
            m2_column = "M2_PZ"
        
        # Build update query dynamically based on which values are provided
        update_fields = []
        update_params = {"ek_id": params.ek_id}
        
        if params.m1 is not None:
            update_fields.append(f"{m1_column} = :m1")
            update_params["m1"] = params.m1
        
        if params.m2 is not None:
            update_fields.append(f"{m2_column} = :m2")
            update_params["m2"] = params.m2
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="At least one of m1 or m2 must be provided")
        
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": params.ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {params.ek_id} not found")
        
        # Perform the update
        update_query = text(f"""
            UPDATE SRTN_EK_SEISM_DATA 
            SET {', '.join(update_fields)}
            WHERE EK_ID = :ek_id
        """)
        
        print(f"Executing update query: {update_query}")
        print(f"Parameters: {update_params}")
        
        result = db.execute(update_query, update_params)
        
        # Check if the update was successful
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No rows updated for EK_ID {params.ek_id}")
        
        # Commit the transaction
        db.commit()
        
        # Prepare response
        updated_fields = {}
        if params.m1 is not None:
            updated_fields[m1_column] = params.m1
        if params.m2 is not None:
            updated_fields[m2_column] = params.m2
        
        return SaveAnalysisResultResponse(
            success=True,
            message=f"Successfully updated analysis results for EK_ID {params.ek_id} ({params.spectrum_type})",
            updated_fields=updated_fields
        )
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error saving analysis result: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving analysis result: {str(e)}")

@app.post("/api/save-stress-inputs", response_model=SaveStressInputsResponse)
async def save_stress_inputs(
    db: DbSessionDep,
    params: SaveStressInputsParams = Body(...)
):
    """
    Save stress input values to SRTN_EK_SEISM_DATA table.
    
    This endpoint updates the SRTN_EK_SEISM_DATA table with stress input values:
    - SIGMA_DOP (σ)
    - HCLPF 
    - SIGMA_1 (σ₁)
    - SIGMA_2 (σ₂)
    - SIGMA_S_1 ((σ₁)₁)
    - SIGMA_S_2 ((σ₁)₂)
    - SIGMA_S_S1 ((σ₁)s₁)
    - SIGMA_S_S2 ((σ₂)s₂)
    """
    try:
        # Build update query dynamically based on which values are provided
        update_fields = []
        update_params = {"ek_id": params.ek_id}
        
        # Map parameter names to database column names
        field_mapping = {
            # Общие характеристики
            "sigma_dop": "SIGMA_DOP",
            "hclpf": "HCLPF",
            "sigma_1": "SIGMA_1",
            "sigma_2": "SIGMA_2",
            # Поля для ПЗ
            "sigma_1_1_pz": "SIGMA_S_1_PZ",
            "sigma_1_2_pz": "SIGMA_S_2_PZ",
            "sigma_1_s1_pz": "SIGMA_S_S1_PZ",
            "sigma_2_s2_pz": "SIGMA_S_S2_PZ",
            # Поля для МРЗ
            "sigma_1_1_mrz": "SIGMA_S_1_MRZ", 
            "sigma_1_2_mrz": "SIGMA_S_2_MRZ",
            "sigma_1_s1_mrz": "SIGMA_S_S1_MRZ",
            "sigma_2_s2_mrz": "SIGMA_S_S2_MRZ"
        }
        
        # Add all fields (including null values to clear them)
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            update_fields.append(f"{db_column} = :{param_name}")
            update_params[param_name] = param_value
        
        # Clear previous calculation results when saving new input data
        # but keep M1 values as they come from spectral analysis, not stress inputs
        update_fields.extend([
            "SIGMA_S_ALT_1_PZ = NULL",
            "SIGMA_S_ALT_2_PZ = NULL", 
            "SIGMA_S_ALT_1_MRZ = NULL",
            "SIGMA_S_ALT_2_MRZ = NULL"
        ])
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="At least one stress input value must be provided")
        
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": params.ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {params.ek_id} not found")
        
        # Perform the update
        update_query = text(f"""
            UPDATE SRTN_EK_SEISM_DATA 
            SET {', '.join(update_fields)}
            WHERE EK_ID = :ek_id
        """)
        
        print(f"Executing stress inputs update query: {update_query}")
        print(f"Parameters: {update_params}")
        
        result = db.execute(update_query, update_params)
        
        # Check if the update was successful
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No rows updated for EK_ID {params.ek_id}")
        
        # Commit the transaction
        db.commit()
        
        # Prepare response with updated fields
        updated_fields = {}
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            if param_value is not None:
                updated_fields[db_column] = param_value
        
        return SaveStressInputsResponse(
            success=True,
            message=f"Successfully updated stress inputs for EK_ID {params.ek_id}",
            updated_fields=updated_fields
        )
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error saving stress inputs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving stress inputs: {str(e)}")

@app.get("/api/get-stress-inputs")
async def get_stress_inputs(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element")
):
    """
    Get stress input values from SRTN_EK_SEISM_DATA table.
    
    This endpoint retrieves stress input values for a specific element:
    - SIGMA_DOP (σ)
    - HCLPF 
    - SIGMA_1 (σ₁)
    - SIGMA_2 (σ₂)
    - SIGMA_S_1 ((σ₁)₁)
    - SIGMA_S_2 ((σ₁)₂)
    - SIGMA_S_S1 ((σ₁)s₁)
    - SIGMA_S_S2 ((σ₂)s₂)
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get stress input values
        stress_query = text("""
            SELECT SIGMA_DOP, HCLPF, SIGMA_1, SIGMA_2, 
                   SIGMA_S_1_PZ, SIGMA_S_2_PZ, SIGMA_S_S1_PZ, SIGMA_S_S2_PZ,
                   SIGMA_S_1_MRZ, SIGMA_S_2_MRZ, SIGMA_S_S1_MRZ, SIGMA_S_S2_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(stress_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Map database columns to response
        stress_values = {
            "SIGMA_DOP": row[0],
            "HCLPF": row[1],
            "SIGMA_1": row[2],
            "SIGMA_2": row[3],
            "SIGMA_S_1_PZ": row[4],
            "SIGMA_S_2_PZ": row[5], 
            "SIGMA_S_S1_PZ": row[6],
            "SIGMA_S_S2_PZ": row[7],
            "SIGMA_S_1_MRZ": row[8],
            "SIGMA_S_2_MRZ": row[9],
            "SIGMA_S_S1_MRZ": row[10],
            "SIGMA_S_S2_MRZ": row[11]
        }
        
        return {
            "success": True,
            "ek_id": ek_id,
            "stress_values": stress_values
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting stress inputs: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving stress inputs: {str(e)}")

@app.get("/api/get-analysis-results")
async def get_analysis_results(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element")
):
    """
    Get analysis results (M1, M2 values) from SRTN_EK_SEISM_DATA table.
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get analysis results
        results_query = text("""
            SELECT M1_PZ, M2_PZ, M1_MRZ, M2_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(results_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Parse results
        analysis_values = {}
        
        if row[0] is not None:
            analysis_values["M1_PZ"] = row[0]
        if row[1] is not None:
            analysis_values["M2_PZ"] = row[1]
        if row[2] is not None:
            analysis_values["M1_MRZ"] = row[2]
        if row[3] is not None:
            analysis_values["M2_MRZ"] = row[3]
        
        return {
            "success": True,
            "ek_id": ek_id,
            "analysis_values": analysis_values
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting analysis results: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving analysis results: {str(e)}")

@app.get("/api/get-calculation-results")
async def get_calculation_results(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element")
):
    """
    Get calculation results (sigma alt values) from SRTN_EK_SEISM_DATA table.
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get calculation results
        results_query = text("""
            SELECT SIGMA_S_ALT_1_PZ, SIGMA_S_ALT_2_PZ, SIGMA_S_ALT_1_MRZ, SIGMA_S_ALT_2_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(results_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Parse results
        calculated_values = {}
        
        if row[0] is not None:
            calculated_values["SIGMA_S_ALT_1_PZ"] = row[0]
        if row[1] is not None:
            calculated_values["SIGMA_S_ALT_2_PZ"] = row[1]
        if row[2] is not None:
            calculated_values["SIGMA_S_ALT_1_MRZ"] = row[2]
        if row[3] is not None:
            calculated_values["SIGMA_S_ALT_2_MRZ"] = row[3]
        
        return {
            "success": True,
            "ek_id": ek_id,
            "calculated_values": calculated_values
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting calculation results: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving calculation results: {str(e)}")

@app.get("/api/check-calculation-requirements")
async def check_calculation_requirements(
    db: DbSessionDep,
    ek_id: int = Query(..., description="ID of the element")
):
    """
    Check what data is missing for sigma alt calculations and return detailed information
    """
    try:
        # Get current data
        data_query = text("""
            SELECT 
                -- PZ data
                SIGMA_S_1_PZ, SIGMA_S_2_PZ, SIGMA_S_S1_PZ, SIGMA_S_S2_PZ, M1_PZ,
                -- MRZ data  
                SIGMA_S_1_MRZ, SIGMA_S_2_MRZ, SIGMA_S_S1_MRZ, SIGMA_S_S2_MRZ, M1_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(data_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Extract values
        (sigma_s_1_pz, sigma_s_2_pz, sigma_s_s1_pz, sigma_s_s2_pz, m1_pz,
         sigma_s_1_mrz, sigma_s_2_mrz, sigma_s_s1_mrz, sigma_s_s2_mrz, m1_mrz) = row
        
        requirements = {
            "pz": {
                "sigma_alt_1": {
                    "can_calculate": all(v is not None for v in [sigma_s_1_pz, sigma_s_s1_pz, m1_pz]),
                    "missing_fields": []
                },
                "sigma_alt_2": {
                    "can_calculate": all(v is not None for v in [sigma_s_2_pz, sigma_s_s2_pz, m1_pz]),
                    "missing_fields": []
                }
            },
            "mrz": {
                "sigma_alt_1": {
                    "can_calculate": all(v is not None for v in [sigma_s_1_mrz, sigma_s_s1_mrz, m1_mrz]),
                    "missing_fields": []
                },
                "sigma_alt_2": {
                    "can_calculate": all(v is not None for v in [sigma_s_2_mrz, sigma_s_s2_mrz, m1_mrz]),
                    "missing_fields": []
                }
            }
        }
        
        # Check missing fields for PZ sigma_alt_1
        if sigma_s_1_pz is None:
            requirements["pz"]["sigma_alt_1"]["missing_fields"].append("(σ₁)₁")
        if sigma_s_s1_pz is None:
            requirements["pz"]["sigma_alt_1"]["missing_fields"].append("(σ₁)s₁")
        if m1_pz is None:
            requirements["pz"]["sigma_alt_1"]["missing_fields"].append("M₁")
            
        # Check missing fields for PZ sigma_alt_2
        if sigma_s_2_pz is None:
            requirements["pz"]["sigma_alt_2"]["missing_fields"].append("(σ₂)₂")
        if sigma_s_s2_pz is None:
            requirements["pz"]["sigma_alt_2"]["missing_fields"].append("(σ₂)s₂")
        if m1_pz is None:
            requirements["pz"]["sigma_alt_2"]["missing_fields"].append("M₁")
            
        # Check missing fields for MRZ sigma_alt_1
        if sigma_s_1_mrz is None:
            requirements["mrz"]["sigma_alt_1"]["missing_fields"].append("(σ₁)₁")
        if sigma_s_s1_mrz is None:
            requirements["mrz"]["sigma_alt_1"]["missing_fields"].append("(σ₁)s₁")
        if m1_mrz is None:
            requirements["mrz"]["sigma_alt_1"]["missing_fields"].append("M₁")
            
        # Check missing fields for MRZ sigma_alt_2
        if sigma_s_2_mrz is None:
            requirements["mrz"]["sigma_alt_2"]["missing_fields"].append("(σ₂)₂")
        if sigma_s_s2_mrz is None:
            requirements["mrz"]["sigma_alt_2"]["missing_fields"].append("(σ₂)s₂")
        if m1_mrz is None:
            requirements["mrz"]["sigma_alt_2"]["missing_fields"].append("M₁")
        
        return {
            "success": True,
            "ek_id": ek_id,
            "requirements": requirements
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error checking calculation requirements: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error checking requirements: {str(e)}")

@app.post("/api/calculate-sigma-alt")
async def calculate_sigma_alt(
    db: DbSessionDep,
    ek_id: int = Body(..., embed=True)
):
    """
    Calculate sigma alt (σ*) values using the formulas:
    (σs)₁* = (σs)₁ + (σs)s₁ · (m₁ - 1)
    (σs)₂* = (σs)₂ + (σs)s₂ · (m₁ - 1)
    
    Results are saved to SIGMA_S_ALT_1_PZ, SIGMA_S_ALT_2_PZ, SIGMA_S_ALT_1_MRZ, SIGMA_S_ALT_2_MRZ
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get current data for calculations
        data_query = text("""
            SELECT 
                -- PZ data
                SIGMA_S_1_PZ, SIGMA_S_2_PZ, SIGMA_S_S1_PZ, SIGMA_S_S2_PZ,
                M1_PZ,
                -- MRZ data  
                SIGMA_S_1_MRZ, SIGMA_S_2_MRZ, SIGMA_S_S1_MRZ, SIGMA_S_S2_MRZ,
                M1_MRZ,
                -- Also get alt values to see if they exist
                SIGMA_S_ALT_1_PZ, SIGMA_S_ALT_2_PZ, SIGMA_S_ALT_1_MRZ, SIGMA_S_ALT_2_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(data_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Extract values
        (sigma_s_1_pz, sigma_s_2_pz, sigma_s_s1_pz, sigma_s_s2_pz, m1_pz,
         sigma_s_1_mrz, sigma_s_2_mrz, sigma_s_s1_mrz, sigma_s_s2_mrz, m1_mrz,
         existing_alt_1_pz, existing_alt_2_pz, existing_alt_1_mrz, existing_alt_2_mrz) = row
        
        print(f"=== CALCULATION DATA FOR EK_ID {ek_id} ===")
        print(f"PZ Data:")
        print(f"  sigma_s_1_pz = {sigma_s_1_pz}")
        print(f"  sigma_s_2_pz = {sigma_s_2_pz}")
        print(f"  sigma_s_s1_pz = {sigma_s_s1_pz}")
        print(f"  sigma_s_s2_pz = {sigma_s_s2_pz}")
        print(f"  m1_pz = {m1_pz}")
        print(f"MRZ Data:")
        print(f"  sigma_s_1_mrz = {sigma_s_1_mrz}")
        print(f"  sigma_s_2_mrz = {sigma_s_2_mrz}")
        print(f"  sigma_s_s1_mrz = {sigma_s_s1_mrz}")
        print(f"  sigma_s_s2_mrz = {sigma_s_s2_mrz}")
        print(f"  m1_mrz = {m1_mrz}")
        print(f"Existing ALT values: PZ({existing_alt_1_pz}, {existing_alt_2_pz}), MRZ({existing_alt_1_mrz}, {existing_alt_2_mrz})")
        print(f"============================================")
        
        update_fields = []
        update_params = {"ek_id": ek_id}
        calculated_values = {}
        
        # Calculate for PZ if data is available
        print(f"Checking PZ calculation requirements:")
        print(f"  sigma_s_1_pz={sigma_s_1_pz}, sigma_s_s1_pz={sigma_s_s1_pz}, m1_pz={m1_pz}")
        if all(v is not None for v in [sigma_s_1_pz, sigma_s_s1_pz, m1_pz]):
            sigma_alt_1_pz = sigma_s_1_pz + sigma_s_s1_pz * (m1_pz - 1)
            update_fields.append("SIGMA_S_ALT_1_PZ = :sigma_alt_1_pz")
            update_params["sigma_alt_1_pz"] = sigma_alt_1_pz
            calculated_values["SIGMA_S_ALT_1_PZ"] = sigma_alt_1_pz
            print(f"  ✓ Calculated sigma_alt_1_pz = {sigma_alt_1_pz}")
        else:
            print(f"  ✗ Cannot calculate sigma_alt_1_pz - missing required values")
        
        print(f"  sigma_s_2_pz={sigma_s_2_pz}, sigma_s_s2_pz={sigma_s_s2_pz}, m1_pz={m1_pz}")
        if all(v is not None for v in [sigma_s_2_pz, sigma_s_s2_pz, m1_pz]):
            sigma_alt_2_pz = sigma_s_2_pz + sigma_s_s2_pz * (m1_pz - 1)
            update_fields.append("SIGMA_S_ALT_2_PZ = :sigma_alt_2_pz")
            update_params["sigma_alt_2_pz"] = sigma_alt_2_pz
            calculated_values["SIGMA_S_ALT_2_PZ"] = sigma_alt_2_pz
            print(f"  ✓ Calculated sigma_alt_2_pz = {sigma_alt_2_pz}")
        else:
            print(f"  ✗ Cannot calculate sigma_alt_2_pz - missing required values")
        
        # Calculate for MRZ if data is available
        print(f"Checking MRZ calculation requirements:")
        print(f"  sigma_s_1_mrz={sigma_s_1_mrz}, sigma_s_s1_mrz={sigma_s_s1_mrz}, m1_mrz={m1_mrz}")
        if all(v is not None for v in [sigma_s_1_mrz, sigma_s_s1_mrz, m1_mrz]):
            sigma_alt_1_mrz = sigma_s_1_mrz + sigma_s_s1_mrz * (m1_mrz - 1)
            update_fields.append("SIGMA_S_ALT_1_MRZ = :sigma_alt_1_mrz")
            update_params["sigma_alt_1_mrz"] = sigma_alt_1_mrz
            calculated_values["SIGMA_S_ALT_1_MRZ"] = sigma_alt_1_mrz
            print(f"  ✓ Calculated sigma_alt_1_mrz = {sigma_alt_1_mrz}")
        else:
            print(f"  ✗ Cannot calculate sigma_alt_1_mrz - missing required values")
        
        print(f"  sigma_s_2_mrz={sigma_s_2_mrz}, sigma_s_s2_mrz={sigma_s_s2_mrz}, m1_mrz={m1_mrz}")
        if all(v is not None for v in [sigma_s_2_mrz, sigma_s_s2_mrz, m1_mrz]):
            sigma_alt_2_mrz = sigma_s_2_mrz + sigma_s_s2_mrz * (m1_mrz - 1)
            update_fields.append("SIGMA_S_ALT_2_MRZ = :sigma_alt_2_mrz")
            update_params["sigma_alt_2_mrz"] = sigma_alt_2_mrz
            calculated_values["SIGMA_S_ALT_2_MRZ"] = sigma_alt_2_mrz
            print(f"  ✓ Calculated sigma_alt_2_mrz = {sigma_alt_2_mrz}")
        else:
            print(f"  ✗ Cannot calculate sigma_alt_2_mrz - missing required values")
        
        if not update_fields:
            print("No calculations performed - insufficient data for both PZ and MRZ")
            return {
                "success": True,
                "message": "No calculations performed - insufficient data for both PZ and MRZ", 
                "calculated_values": {}
            }
        
        # Perform the update
        update_query = text(f"""
            UPDATE SRTN_EK_SEISM_DATA 
            SET {', '.join(update_fields)}
            WHERE EK_ID = :ek_id
        """)
        
        print(f"Executing sigma alt calculation query: {update_query}")
        print(f"Parameters: {update_params}")
        
        result = db.execute(update_query, update_params)
        
        # Check if the update was successful
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No rows updated for EK_ID {ek_id}")
        
        # Commit the transaction
        db.commit()
        
        # Verify that the values were actually saved
        verify_query = text("""
            SELECT SIGMA_S_ALT_1_PZ, SIGMA_S_ALT_2_PZ, SIGMA_S_ALT_1_MRZ, SIGMA_S_ALT_2_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        verify_result = db.execute(verify_query, {"ek_id": ek_id})
        verify_row = verify_result.fetchone()
        
        if verify_row:
            print(f"Verification - saved ALT values: PZ({verify_row[0]}, {verify_row[1]}), MRZ({verify_row[2]}, {verify_row[3]})")
        
        return {
            "success": True,
            "message": f"Successfully calculated sigma alt values for EK_ID {ek_id}",
            "calculated_values": calculated_values
        }
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error calculating sigma alt: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error calculating sigma alt: {str(e)}")

@app.post("/api/save-k-results", response_model=SaveKResultsResponse)
async def save_k_results(
    db: DbSessionDep,
    params: SaveKResultsParams = Body(...)
):
    """
    Save K coefficient results to SRTN_EK_SEISM_DATA table.
    
    This endpoint updates the SRTN_EK_SEISM_DATA table with K coefficient values:
    - K1_PZ, K2_PZ, K_MIN_PZ for ПЗ
    - K1_MRZ, K2_MRZ, K_MIN_MRZ for МРЗ  
    - SEISMIC_CATEGORY_PZ for seismic category
    """
    try:
        # Build update query dynamically based on which values are provided
        update_fields = []
        update_params = {"ek_id": params.ek_id}
        
        # Map parameter names to database column names - сохраняем k1 в K1 поля
        field_mapping = {
            # Сохраняем k1 значения в K1 поля
            "k1_pz": "K1_PZ",      # k1 для ПЗ в K1_PZ
            "k1_mrz": "K1_MRZ"     # k1 для МРЗ в K1_MRZ
        }
        
        # Add all fields (including null values to clear them)
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            update_fields.append(f"{db_column} = :{param_name}")
            update_params[param_name] = param_value
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="At least one K coefficient value must be provided")
        
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": params.ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {params.ek_id} not found")
        
        # Perform the update
        update_query = text(f"""
            UPDATE SRTN_EK_SEISM_DATA 
            SET {', '.join(update_fields)}
            WHERE EK_ID = :ek_id
        """)
        
        print(f"Executing K results update query: {update_query}")
        print(f"Parameters: {update_params}")
        
        result = db.execute(update_query, update_params)
        
        # Check if the update was successful
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No rows updated for EK_ID {params.ek_id}")
        
        # Commit the transaction
        db.commit()
        
        # Prepare response with updated fields
        updated_fields = {}
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            if param_value is not None:
                updated_fields[db_column] = param_value
        
        return SaveKResultsResponse(
            success=True,
            message=f"Successfully updated K coefficient results for EK_ID {params.ek_id}",
            updated_fields=updated_fields
        )
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error saving K results: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving K results: {str(e)}")

@app.get("/api/get-k-results/{ek_id}")
async def get_k_results(
    db: DbSessionDep,
    ek_id: int
):
    """
    Get K coefficient results from SRTN_EK_SEISM_DATA table.
    
    This endpoint retrieves K coefficient values for a specific element:
    - K1_PZ, K2_PZ, K_MIN_PZ for ПЗ
    - K1_MRZ, K2_MRZ, K_MIN_MRZ for МРЗ
    - SEISMIC_CATEGORY_PZ for seismic category
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get K coefficient values - только K1_PZ, K1_MRZ
        k_query = text("""
            SELECT K1_PZ, K1_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(k_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Check if any K values exist
        has_k_data = any(value is not None for value in row)
        
        # Map database columns to response - возвращаем правильную структуру
        k_values = {
            "k1_pz": row[0],      # K1_PZ содержит k1 для ПЗ
            "k2_pz": None,        # K2_PZ не используется
            "k_min_pz": row[0],   # K1_PZ - минимальное значение для ПЗ
            "seismic_category_pz": None,  # Не сохраняем в БД
            "k1_mrz": row[1],     # K1_MRZ содержит k1 для МРЗ
            "k2_mrz": None,       # K2_MRZ не используется
            "k_min_mrz": row[1],  # K1_MRZ - минимальное значение для МРЗ
            "calculated": has_k_data
        }
        
        return {
            "success": True,
            "ek_id": ek_id,
            **k_values
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting K results: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving K results: {str(e)}")

class LoadAnalysisParams(BaseModel):
    element_id: int
    material: str | None = None
    doc_code_analytics: str | None = None
    doc_code_operation: str | None = None
    p1_pz: float | None = None
    temp1_pz: float | None = None
    p2_pz: float | None = None
    temp2_pz: float | None = None
    sigma_dop_a_pz: float | None = None
    ratio_e_pz: float | None = None
    p1_mrz: float | None = None
    temp1_mrz: float | None = None
    p2_mrz: float | None = None
    temp2_mrz: float | None = None
    sigma_dop_a_mrz: float | None = None
    ratio_e_mrz: float | None = None

@app.post("/api/save-load-analysis-params")
async def save_load_analysis_params(
    db: DbSessionDep,
    params: LoadAnalysisParams = Body(...)
):
    """
    Save load analysis parameters to SRTN_EK_SEISM_DATA table.
    
    This endpoint updates the SRTN_EK_SEISM_DATA table with load analysis parameters:
    - Material name (MAT_NAME)
    - Document codes (DOC_1, DOC_2)
    - Load analysis parameters for PZ and MRZ
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": params.element_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {params.element_id} not found")
        
        # Build update query dynamically based on which values are provided
        update_fields = []
        update_params = {"ek_id": params.element_id}
        
        # Map parameter names to database column names
        field_mapping = {
            # Общие параметры
            "material": "MAT_NAME",
            "doc_code_analytics": "DOC_1", 
            "doc_code_operation": "DOC_2",
            # Параметры ПЗ
            "p1_pz": "P1_PZ",
            "temp1_pz": "TEMP1_PZ", 
            "p2_pz": "P2_PZ",
            "temp2_pz": "TEMP2_PZ",
            "sigma_dop_a_pz": "SIGMA_DOP_A_PZ",
            "ratio_e_pz": "RATIO_E_PZ",
            # Параметры МРЗ
            "p1_mrz": "P1_MRZ",
            "temp1_mrz": "TEMP1_MRZ",
            "p2_mrz": "P2_MRZ", 
            "temp2_mrz": "TEMP2_MRZ",
            "sigma_dop_a_mrz": "SIGMA_DOP_A_MRZ",
            "ratio_e_mrz": "RATIO_E_MRZ"
        }
        
        # Add all fields (including null values to clear them)
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            update_fields.append(f"{db_column} = :{param_name}")
            update_params[param_name] = param_value
        
        if not update_fields:
            raise HTTPException(status_code=400, detail="At least one load analysis parameter must be provided")
        
        # Perform the update
        update_query = text(f"""
            UPDATE SRTN_EK_SEISM_DATA 
            SET {', '.join(update_fields)}
            WHERE EK_ID = :ek_id
        """)
        
        print(f"Executing load analysis params update query: {update_query}")
        print(f"Parameters: {update_params}")
        
        result = db.execute(update_query, update_params)
        
        # Check if the update was successful
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail=f"No rows updated for EK_ID {params.element_id}")
        
        # Commit the transaction
        db.commit()
        
        # Prepare response with updated fields
        updated_fields = {}
        for param_name, db_column in field_mapping.items():
            param_value = getattr(params, param_name)
            if param_value is not None:
                updated_fields[db_column] = param_value
        
        return {
            "success": True,
            "message": f"Successfully saved load analysis parameters for EK_ID {params.element_id}",
            "updated_fields": updated_fields
        }
        
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        print(f"Error saving load analysis parameters: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error saving load analysis parameters: {str(e)}")

@app.get("/api/get-load-analysis-params/{ek_id}")
async def get_load_analysis_params(
    db: DbSessionDep,
    ek_id: int
):
    """
    Get load analysis parameters from SRTN_EK_SEISM_DATA table.
    
    This endpoint retrieves load analysis parameters for a specific element:
    - Material name (MAT_NAME)
    - Document codes (DOC_1, DOC_2)
    - Load analysis parameters for PZ and MRZ
    """
    try:
        # Check if the EK_ID exists
        check_query = text("SELECT COUNT(*) FROM SRTN_EK_SEISM_DATA WHERE EK_ID = :ek_id")
        check_result = db.execute(check_query, {"ek_id": ek_id})
        if check_result.scalar() == 0:
            raise HTTPException(status_code=404, detail=f"Element with EK_ID {ek_id} not found")
        
        # Get load analysis parameters
        params_query = text("""
            SELECT 
                MAT_NAME, DOC_1, DOC_2,
                P1_PZ, TEMP1_PZ, P2_PZ, TEMP2_PZ, SIGMA_DOP_A_PZ, RATIO_E_PZ,
                P1_MRZ, TEMP1_MRZ, P2_MRZ, TEMP2_MRZ, SIGMA_DOP_A_MRZ, RATIO_E_MRZ
            FROM SRTN_EK_SEISM_DATA 
            WHERE EK_ID = :ek_id
        """)
        
        result = db.execute(params_query, {"ek_id": ek_id})
        row = result.fetchone()
        
        if not row:
            raise HTTPException(status_code=404, detail=f"No data found for EK_ID {ek_id}")
        
        # Map database columns to response
        load_params = {
            "material": row[0],
            "doc_code_analytics": row[1],
            "doc_code_operation": row[2],
            "p1_pz": row[3],
            "temp1_pz": row[4],
            "p2_pz": row[5],
            "temp2_pz": row[6],
            "sigma_dop_a_pz": row[7],
            "ratio_e_pz": row[8],
            "p1_mrz": row[9],
            "temp1_mrz": row[10],
            "p2_mrz": row[11],
            "temp2_mrz": row[12],
            "sigma_dop_a_mrz": row[13],
            "ratio_e_mrz": row[14]
        }
        
        return {
            "success": True,
            "ek_id": ek_id,
            "load_params": load_params
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error getting load analysis parameters: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error retrieving load analysis parameters: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    
    if settings.dev:
        uvicorn.run("main:app", host="127.0.0.1", port=8000, reload=True)
    else:
        uvicorn.run(app, host="0.0.0.0", port=8000)
