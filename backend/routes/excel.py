import io
import re

import openpyxl
from fastapi import APIRouter, File, Form, HTTPException, UploadFile


router = APIRouter(prefix="/api", tags=["excel"])


@router.post("/analyze-excel")
async def analyze_excel(
    file: UploadFile = File(...),
    filter_percentage_only: bool = False,
):
    if not file.filename.endswith((".xls", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xls, .xlsx, .xlsm)")

    try:
        contents = await file.read()
        file_object = io.BytesIO(contents)

        workbook = openpyxl.load_workbook(file_object, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames

        percentage_pattern = re.compile(r'^[0-9]+([.,][0-9]+)?%$')

        sheets_info = []
        for sheet_name in sheet_names:
            if filter_percentage_only and not percentage_pattern.match(sheet_name.strip()):
                continue

            sheet = workbook[sheet_name]
            try:
                max_row = sheet.max_row
                max_col = sheet.max_column
                if max_row > 0 and max_col > 0:
                    sheets_info.append({"name": sheet_name, "rows": max_row, "columns": max_col})
                else:
                    sheets_info.append({"name": sheet_name, "rows": 0, "columns": 0})
            except Exception as sheet_error:
                sheets_info.append({"name": sheet_name, "rows": 0, "columns": 0, "error": str(sheet_error)})

        workbook.close()
        file_object.close()

        return {"sheets": sheets_info}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error analyzing Excel file: {str(e)}")


@router.post("/extract-sheet-data")
async def extract_sheet_data(
    file: UploadFile = File(...),
    sheet_name: str = Form(...),
):
    if not file.filename.endswith((".xls", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xls, .xlsx, .xlsm)")

    try:
        contents = await file.read()
        file_object = io.BytesIO(contents)

        workbook = openpyxl.load_workbook(file_object, read_only=True, data_only=True)
        if sheet_name not in workbook.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{sheet_name}' not found in the Excel file")

        sheet = workbook[sheet_name]
        headers = []
        max_col = sheet.max_column
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value))

        result = {}

        for col_idx, header in enumerate(headers, start=1):
            column_data = []
            for row_idx in range(2, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    column_data.append(str(cell_value))
            col_key = f"column_{col_idx}" if not header.strip() else header
            result[col_key] = column_data

        workbook.close()
        file_object.close()
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error extracting sheet data: {str(e)}")


